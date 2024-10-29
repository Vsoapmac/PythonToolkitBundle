import warnings
import xlwings as xw
from pathlib import Path
from xlrd import open_workbook
from xlutils.copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


def get_column_letter(column_index: int) -> str:
    """通过列索引获取列名

    Args:
        column_index (int): 第几列, 用1、2、3等表示(从1开始)

    Returns:
        str: 列名, 用A、B、C等表示
    """
    if column_index <= 0:
        raise ValueError("Column number must be positive.")
    result = ""
    while column_index > 0:
        column_index -= 1  # 转换为从0开始的索引
        remainder = column_index % 26
        result = chr(65 + remainder) + result  # 65是ASCII码中'A'的值
        column_index //= 26
    return result

def get_column_index(column: str) -> int:
    """通过列名获取列索引

    Args:
        column (str): 列名, 用A、B、C等表示

    Returns:
        int: 列索引, 用1、2、3等表示(从1开始)
    """
    result = 0
    for i, letter in enumerate(reversed(column.upper())):
        if 'A' <= letter <= 'Z':
            result += (ord(letter) - ord('A') + 1) * (26 ** i)
        else:
            raise ValueError("Invalid column letter.")
    return result

class OpenpyxlAdapter:
    
    def __init__(self, book, sheet, excel_path, sheet_name):
        self.book = book
        self.sheet = sheet
        self.excel_path = excel_path
        self.sheet_name = sheet_name

    def close(self):
        """关闭引擎"""
        self.book = None
        self.sheet = None
        self.excel_path = None
        self.sheet_name = None
    
    def quit(self):
        """完全退出引擎, 会先保存表格后退出"""
        self.book.save(self.excel_path)
        self.book = None
        self.sheet = None
        self.excel_path = None
        self.sheet_name = None
    
    def save(self):
        """保存表格"""
        self.book.save(self.excel_path)
    
    def get_value(self, column: str, row: int):
        """获取单元格数据

        Args:
            column (str): 列, 用A、B、C等表示
            row (int): 行, 用1、2、3等表示(从1开始)

        Returns:
            单元格数据
        """
        return self.sheet[f"{column}{row}"].value

    def get_row_values(self, row: int) -> list:
        """获取某行数据

        Args:
            row (int): 行, 用1、2、3等表示(从1开始)

        Returns:
            list: 行数据
        """
        row_values = []
        none_counter = 0
        for cell in self.sheet[row]:
            row_values.append(cell.value)
            if cell.value is not None:
                none_counter = 0
            else:
                none_counter += 1
            if none_counter >= 5:
                break
        return row_values[0: len(row_values)-5] if none_counter >= 5 else row_values
        
    def get_col_values(self, column: str) -> list:
        """获取某列数据

        Args:
            column (str): 列, 用A、B、C等表示

        Returns:
            list: 列数据
        """
        col_values = []
        none_counter = 0
        for cell in self.sheet[column]:
            col_values.append(cell.value)
            if cell.value:
                none_counter = 0
            else:
                none_counter += 1
            if none_counter >= 5:
                break
        return col_values[0: len(col_values)-5] if none_counter >= 5 else col_values
        
    def insert_value(self, column: str, row: int, value, auto_save: bool=True):
        """插入数据，也可以更新某一单元格的数据

        Args:
            column (str): 列, 用A、B、C等表示
            row (int): 行, 用1、2、3等表示(从1开始)
            value (_type_): 要插入/更新的数据
            auto_save (bool, optional): 自动保存开关. Defaults to True.
        """
        self.sheet[f"{column}{row}"] = value
        if auto_save:
            self.book.save(self.excel_path)
    
    def append_value(self, column: str, value, start_row: int=1, auto_save: bool=True):
        """在表末尾附加数据

        Args:
            column (str): 列, 用A、B、C等表示
            value (_type_): 要添加的数据
            start_row (int, optional): 从哪一行开始检测. Defaults to 1.
            auto_save (bool, optional): 自动保存开关. Defaults to True.
        """
        current_row = 1
        for cell in self.sheet[column]:
            if cell.value is None and start_row <= current_row:
                cell.value = value
                break
            current_row += 1
        if auto_save:
            self.book.save(self.excel_path)
    
    def insert_values(self, column: str, values: list, insert_mode: str="append", start_row: int=1, auto_save: bool=True):
        """插入多个数据，也可以更新某一单元格的数据

        Args:
            column (str): 列, 用A、B、C等表示
            values (list): 要添加的多个数据
            insert_mode (str, optional): 程序插入数据所遵循的模式, append和insert, append会在表后附加数据, insert会直接覆盖现有的数据进行插入. Defaults to "append".
            start_row (int, optional): append模式时, 从哪一行开始检测; insert模式时, 从哪一行开始插入. Defaults to 1.
            auto_save (bool, optional): 自动保存开关. Defaults to True.
        """
        if insert_mode == "append":
            current_row = 1
            i = 0
            for cell in self.sheet[column]:
                if cell.value is None and start_row <= current_row:
                    cell.value = values[i]
                    i += 1
                if i == len(values):
                    break
                current_row += 1
        elif insert_mode == "insert":
            row = start_row
            for value in values:
                self.sheet[f"{column}{row}"] = value
                row += 1
        if auto_save:
            self.book.save(self.excel_path)
    
    def append_row(self, values: list, auto_save: bool=True):
        """在表末尾附加一行数据

        Args:
            values (list): 要添加的多个数据
            auto_save (bool, optional): 自动保存开关. Defaults to True.
        """
        self.sheet.append(values)
        if auto_save:
            self.book.save(self.excel_path)

    def delete_row(self, row: int, auto_save: bool=True):
        """删除一整行数据

        Args:
            row (int): 行, 用1、2、3等表示(从1开始)
            auto_save (bool, optional): 自动保存开关. Defaults to True.
        """
        self.sheet.delete_rows(row)
        if auto_save:
            self.book.save(self.excel_path)

    def delete_col(self, column: str, auto_save: bool=True):
        """删除一整列数据

        Args:
            column (str): 列, 用A、B、C等表示
            auto_save (bool, optional): 自动保存开关. Defaults to True.
        """
        self.sheet.delete_rows(get_column_index(column))
        if auto_save:
            self.book.save(self.excel_path)
    
    def merge_cells(self, start_column: str, start_row: int, end_column: str, end_row: int):
        """合并单元格

        Args:
            start_column (str): 开始列, 用A、B、C等表示
            start_row (int): 开始行, 用1、2、3等表示(从1开始)
            end_column (str): 结束列, 用A、B、C等表示
            end_row (int): 结束行, 用1、2、3等表示(从1开始)
        """
        self.sheet.merge_cells(f"{start_column}{start_row}:{end_column}{end_row}")
    
    def unmerge_cells(self, start_column: str, start_row: int, end_column: str, end_row: int):
        """解除合并单元格

        Args:
            start_column (str): 开始列, 用A、B、C等表示
            start_row (int): 开始行, 用1、2、3等表示(从1开始)
            end_column (str): 结束列, 用A、B、C等表示
            end_row (int): 结束行, 用1、2、3等表示(从1开始)
        """
        self.sheet.unmerge_cells(f"{start_column}{start_row}:{end_column}{end_row}")
    
    def get_font(self, column: str, row: int):
        """获取单元格字体

        Args:
            column (str): 列, 用A、B、C等表示
            row (int): 行, 用1、2、3等表示(从1开始)

        Returns:
            单元格字体
        """
        return self.sheet.cell(row=row, column=get_column_index(column)).font
    
    def set_font(self, column: str, row: int, name: str=None, size: int=None, bold: bool=None, italic: bool=None, color_code: str=None, auto_save: bool=True):
        """设置单元格字体

        Args:
            column (str): 列, 用A、B、C等表示
            row (int): 行, 用1、2、3等表示(从1开始)
            name (str, optional): 字体. Defaults to None.
            size (int, optional): 字符大小. Defaults to None.
            bold (bool, optional): 粗体. Defaults to None.
            italic (bool, optional): 斜体. Defaults to None.
            color_code (str, optional): 16进制颜色码. Defaults to None.
            auto_save (bool, optional): 自动保存开关. Defaults to True.
        """
        color_code = color_code.replace("#", "")
        font = Font(name=name, size=size, bold=bold, italic=italic, color=color_code)
        self.sheet.cell(row=row, column=get_column_index(column)).font = font
        if auto_save:
            self.book.save(self.excel_path)
            
    def get_color(self, column: str, row: int):
        """获取单元格颜色

        Args:
            column (str): 列, 用A、B、C等表示
            row (int): 行, 用1、2、3等表示(从1开始)

        Returns:
            单元格颜色
        """
        return self.sheet.cell(row=row, column=get_column_index(column)).fill

    def set_color(self, column: str, row: int, color_code: str=None, fill_type: str="solid", auto_save: bool=True):
        """填充单元格背景色

        Args:
            column (str): 列, 用A、B、C等表示
            row (int): 行, 用1、2、3等表示(从1开始)
            color_code (str, optional): 16进制颜色码. Defaults to None.
            fill_type (str, optional): 填充类型. Defaults to solid.
            auto_save (bool, optional): 自动保存开关. Defaults to True.
        """
        color_code = color_code.replace("#", "")
        fill = PatternFill(start_color=color_code, end_color=color_code, fill_type=fill_type)
        self.sheet.cell(row=row, column=get_column_index(column)).fill = fill
        if auto_save:
            self.book.save(self.excel_path)

    def get_border(self, column: str, row: int):
        """获取单元格边框

        Args:
            column (str): 列, 用A、B、C等表示
            row (int): 行, 用1、2、3等表示(从1开始)

        Returns:
            单元格边框
        """
        return self.sheet.cell(row=row, column=get_column_index(column)).border
    
    def set_border(self, column: str, row: int, color_code: str="#000000", border_style: str="thin", auto_save: bool=True):
        """设置单元格边框

        Args:
            column (str): 列, 用A、B、C等表示
            row (int): 行, 用1、2、3等表示(从1开始)
            color_code (str, optional): 16进制颜色码. Defaults to "#000000".
            border_style (str, optional): 边框类型. Defaults to "thin".
            auto_save (bool, optional): 自动保存开关. Defaults to True.
        """
        color_code = color_code.replace("#", "")
        border = Border(left=Side(border_style=border_style, color=color_code),
                right=Side(border_style=border_style, color=color_code),
                top=Side(border_style=border_style, color=color_code),
                bottom=Side(border_style=border_style, color=color_code))
        self.sheet.cell(row=row, column=get_column_index(column)).border = border
        if auto_save:
            self.book.save(self.excel_path)
            
    def get_align(self, column: str, row: int):
        """获取单元格对齐方式

        Args:
            column (str): 列, 用A、B、C等表示
            row (int): 行, 用1、2、3等表示(从1开始)

        Returns:
            单元格对齐方式
        """
        return self.sheet.cell(row=row, column=get_column_index(column)).alignment
    
    def set_align(self, column: str, row: int, horizontal: str="center", vertical: str="center", wrap_text: bool=True, auto_save: bool=True):
        """设置单元格对齐方式

        Args:
            column (str): 列, 用A、B、C等表示
            row (int): 行, 用1、2、3等表示(从1开始)
            horizontal (str, optional): 垂直对齐方式. Defaults to "center".
            vertical (str, optional): 水平对齐方式. Defaults to "center".
            wrap_text (bool, optional): 超过文本框后自动换行. Defaults to True.
            auto_save (bool, optional): 自动保存开关. Defaults to True.
        """
        alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)
        self.sheet.cell(row=row, column=get_column_index(column)).alignment = alignment
        if auto_save:
            self.book.save(self.excel_path)

class XlutilsAdapter:
    
    def __init__(self, book, sheet, excel_path, sheet_name):
        self.book = book
        self.sheet = sheet
        self.excel_path = excel_path
        self.sheet_name = sheet_name
    
    def close(self):
        """关闭引擎"""
        self.book = None
        self.sheet = None
        self.excel_path = None
        self.sheet_name = None
    
    def quit(self):
        """完全退出引擎, 会先保存表格后退出"""
        self.book = None
        self.sheet = None
        self.excel_path = None
        self.sheet_name = None
    
    def get_value(self, column: str, row: int):
        """获取单元格数据

        Args:
            column (str): 列, 用A、B、C等表示
            row (int): 行, 用1、2、3等表示(从1开始)

        Returns:
            单元格数据
        """
        return self.sheet.cell(row-1, get_column_index(column)-1).value

    def get_row_values(self, row: int) -> list:
        """获取某行数据

        Args:
            row (int): 行, 用1、2、3等表示(从1开始)

        Returns:
            list: 行数据
        """
        return self.sheet.row_values(row-1)
        
    def get_col_values(self, column: str) -> list:
        """获取某列数据

        Args:
            column (str): 列, 用A、B、C等表示

        Returns:
            list: 列数据
        """
        return self.sheet.col_values(get_column_index(column)-1)
    
    def insert_value(self, column: str, row: int, value, auto_save: bool=True):
        """插入数据，也可以更新某一单元格的数据

        Args:
            column (str): 列, 用A、B、C等表示
            row (int): 行, 用1、2、3等表示(从1开始)
            value (_type_): 要插入/更新的数据
            auto_save (bool, optional): 自动保存开关, 若关闭则会在当前路径保存一份新的文件. Defaults to True.
        """
        warnings.warn("xlutils will lose the original file format after saving the file in xls format, so please save it carefully")
        new_workbook = copy(self.book)
        worksheet = new_workbook.get_sheet(self.sheet_name) # 通过sheet名称获取
        worksheet.write(row-1, get_column_index(column)-1, value)
        if auto_save:
            new_workbook.save(self.excel_path)
            self.book = open_workbook(self.excel_path)
            self.sheet = self.book.sheet_by_name(self.sheet_name)
        else:
            path = Path(self.excel_path)
            parent = path.parent
            filename = path.name.replace(suffix, "")
            suffix = path.suffix
            bck_path = str(Path(parent, f"{filename}_bck{suffix}").resolve())
            new_workbook.save(bck_path)
            self.book = open_workbook(bck_path)
            self.sheet = self.book.sheet_by_name(self.sheet_name)
        
    def append_value(self, column: str, value, auto_save: bool=True):
        """在表末尾附加数据

        Args:
            column (str): 列, 用A、B、C等表示
            value (_type_): 要添加的数据
            auto_save (bool, optional): 自动保存开关, 若关闭则会在当前路径保存一份新的文件. Defaults to True.
        """
        warnings.warn("xlutils will lose the original file format after saving the file in xls format, so please save it carefully")
        new_workbook = copy(self.book)
        worksheet = new_workbook.get_sheet(self.sheet_name) # 通过sheet名称获取
        worksheet.write(self.sheet.nrows, get_column_index(column)-1, value)
        if auto_save:
            new_workbook.save(self.excel_path)
            self.book = open_workbook(self.excel_path)
            self.sheet = self.book.sheet_by_name(self.sheet_name)
        else:
            path = Path(self.excel_path)
            parent = path.parent
            filename = path.name.replace(suffix, "")
            suffix = path.suffix
            bck_path = str(Path(parent, f"{filename}_bck{suffix}").resolve())
            new_workbook.save(bck_path)
            self.book = open_workbook(bck_path)
            self.sheet = self.book.sheet_by_name(self.sheet_name)
        
    def insert_values(self, column: str, values: list, insert_mode: str="append", start_row: int=1, auto_save: bool=True):
        """插入多个数据，也可以更新某一单元格的数据

        Args:
            column (str): 列, 用A、B、C等表示
            values (list): 要添加的多个数据
            insert_mode (str, optional): 程序插入数据所遵循的模式, append和insert, append会在表后附加数据, insert会直接覆盖现有的数据进行插入. Defaults to "append".
            start_row (int, optional): append模式时, 从哪一行开始检测; insert模式时, 从哪一行开始插入. Defaults to 1.
            auto_save (bool, optional): 自动保存开关, 若关闭则会在当前路径保存一份新的文件. Defaults to True.
        """
        warnings.warn("xlutils will lose the original file format after saving the file in xls format, so please save it carefully")
        new_workbook = copy(self.book)
        worksheet = new_workbook.get_sheet(self.sheet_name) # 通过sheet名称获取
        if insert_mode == "append":
            current_row = self.sheet.nrows
            for value in values:
                worksheet.write(current_row, get_column_index(column)-1, value)
                current_row += 1
        elif insert_mode == "insert":
            current_row = start_row
            for value in values:
                worksheet.write(current_row, get_column_index(column)-1, value)
                current_row += 1
        if auto_save:
            new_workbook.save(self.excel_path)
            self.book = open_workbook(self.excel_path)
            self.sheet = self.book.sheet_by_name(self.sheet_name)
        else:
            path = Path(self.excel_path)
            parent = path.parent
            filename = path.name.replace(suffix, "")
            suffix = path.suffix
            bck_path = str(Path(parent, f"{filename}_bck{suffix}").resolve())
            new_workbook.save(bck_path)
            self.book = open_workbook(bck_path)
            self.sheet = self.book.sheet_by_name(self.sheet_name)

    def append_row(self, values: list, auto_save: bool=True):
        """在表末尾附加一行数据

        Args:
            values (list): 要添加的多个数据
            auto_save (bool, optional): 自动保存开关, 若关闭则会在当前路径保存一份新的文件. Defaults to True.
        """
        warnings.warn("xlutils will lose the original file format after saving the file in xls format, so please save it carefully")
        new_workbook = copy(self.book)
        worksheet = new_workbook.get_sheet(self.sheet_name) # 通过sheet名称获取
        current_col = 0
        for value in values:
            worksheet.write(self.sheet.nrows, current_col, value)
            current_col += 1
        if auto_save:
            new_workbook.save(self.excel_path)
            self.book = open_workbook(self.excel_path)
            self.sheet = self.book.sheet_by_name(self.sheet_name)
        else:
            path = Path(self.excel_path)
            parent = path.parent
            filename = path.name.replace(suffix, "")
            suffix = path.suffix
            bck_path = str(Path(parent, f"{filename}_bck{suffix}").resolve())
            new_workbook.save(bck_path)
            self.book = open_workbook(bck_path)
            self.sheet = self.book.sheet_by_name(self.sheet_name)

class XlwingsAdapter:

    def __init__(self, app, book, sheet, excel_path, sheet_name):
        self.app = app
        self.book = book
        self.sheet = sheet
        self.excel_path = excel_path
        self.sheet_name = sheet_name
    
    def close(self):
        """关闭引擎"""
        self.book.close()
        self.app.quit()
        self.app = None
        self.book = None
        self.sheet = None
        self.excel_path = None
        self.sheet_name = None
    
    def quit(self):
        """完全退出引擎, 会先保存表格后退出"""
        self.book.save()
        self.book.close()
        self.app.quit()
        self.app = None
        self.book = None
        self.sheet = None
        self.excel_path = None
        self.sheet_name = None
    
    def save(self):
        """保存表格"""
        self.book.save()
    
    def get_value(self, column: str, row: int):
        """获取单元格数据

        Args:
            column (str): 列, 用A、B、C等表示
            row (int): 行, 用1、2、3等表示(从1开始)

        Returns:
            单元格数据
        """
        return self.sheet[f"{column}{row}"].value
    
    def get_row_values(self, row: int) -> list:
        """获取某行数据

        Args:
            row (int): 行, 用1、2、3等表示(从1开始)

        Returns:
            list: 行数据
        """
        row_data = []
        none_counter = 0
        # 从第一列开始迭代，直到找到一个空单元格为止
        column_letter = 'A'
        for _ in range(16384):
            cell_value = self.sheet.range(f"{column_letter}{row}").value
            row_data.append(cell_value)  # 否则，将值添加到行数据中
            if cell_value:
                none_counter = 0
            elif none_counter >= 5:
                break
            else:
                none_counter += 1
            column_letter = chr(ord(column_letter) + 1)  # 移动到下一列
        return row_data[0: len(row_data)-5] if none_counter >= 5 else row_data
        
    def get_col_values(self, column: str) -> list:
        """获取某列数据

        Args:
            column (str): 列, 用A、B、C等表示

        Returns:
            list: 列数据
        """
        col_value_list = []
        none_counter = 0
        for row in range(1048576):
            value = self.sheet[f"{column}{row+1}"].value
            col_value_list.append(value)
            if value:
                none_counter = 0
            elif none_counter >= 5:
                break
            else:
                none_counter += 1
        return col_value_list[0: len(col_value_list)-5] if none_counter >= 5 else col_value_list
        
    def insert_value(self, column: str, row: int, value, auto_save: bool=True):
        """插入数据，也可以更新某一单元格的数据

        Args:
            column (str): 列, 用A、B、C等表示
            row (int): 行, 用1、2、3等表示(从1开始)
            value (_type_): 要插入/更新的数据
            auto_save (bool, optional): 自动保存开关. Defaults to True.
        """
        self.sheet[f"{column}{row}"].value = value
        if auto_save:
            self.book.save()
    
    def append_value(self, column: str, value, start_row: int=1, auto_save: bool=True):
        """在表末尾附加数据

        Args:
            column (str): 列, 用A、B、C等表示
            value (_type_): 要添加的数据
            start_row (int, optional): 从哪一行开始检测. Defaults to 1.
            auto_save (bool, optional): 自动保存开关. Defaults to True.
        """
        row = start_row
        while row < 1048576:
            col = self.sheet[f"{column}{row}"]
            if col.value == None:
                col.value = value
                return True
            else:
                row += 1
        if auto_save:
            self.book.save()
    
    def insert_values(self, column: str, values: list, insert_mode: str="append", start_row: int=1, auto_save: bool=True):
        """插入多个数据，也可以更新某一单元格的数据

        Args:
            column (str): 列, 用A、B、C等表示
            values (list): 要添加的多个数据
            insert_mode (str, optional): 程序插入数据所遵循的模式, append和insert, append会在表后附加数据, insert会直接覆盖现有的数据进行插入. Defaults to "append".
            start_row (int, optional): append模式时, 从哪一行开始检测; insert模式时, 从哪一行开始插入. Defaults to 1.
            auto_save (bool, optional): 自动保存开关. Defaults to True.
        """
        save_row = start_row
        if insert_mode == "append":
            for value in values:
                while save_row < 1048576:
                    col = self.sheet[f"{column}{save_row}"]
                    if col.value == None:
                        col.value = value
                        save_row += 1
                        break
                    else:
                        save_row += 1
        elif insert_mode == "insert":
            for value in values:
                self.sheet[f"{column}{save_row}"].value = value
                save_row += 1
        if auto_save:
            self.book.save()
    
    def append_row(self, values: list, auto_save: bool=True):
        """在表末尾附加一行数据

        Args:
            values (list): 要添加的多个数据
            auto_save (bool, optional): 自动保存开关. Defaults to True.
        """
        last_row = self.sheet.used_range.last_cell.row
        current_col = 1
        for value in values:
            self.sheet[f"{get_column_letter(current_col)}{last_row}"].value = value
            current_col += 1
        if auto_save:
            self.book.save()

    def delete_row(self, row: int, auto_save: bool=True):
        """删除一整行数据

        Args:
            row (int): 行, 用1、2、3等表示(从1开始)
            auto_save (bool, optional): 自动保存开关. Defaults to True.
        """
        self.sheet.range(f"{row}:{row}").delete()
        if auto_save:
            self.book.save()

    def delete_col(self, column: str, auto_save: bool=True):
        """删除一整列数据

        Args:
            column (str): 列, 用A、B、C等表示
            auto_save (bool, optional): 自动保存开关. Defaults to True.
        """
        self.sheet.range(f"{column}:{column}").delete()
        if auto_save:
            self.book.save()
    
    def get_row_height(self, column: str, row: int) -> float:
        """获取单元格高度

        Args:
            column (str): 列, 用A、B、C等表示
            row (int): 行, 用1、2、3等表示(从1开始)

        Returns:
            float: 单元格高度
        """
        return self.sheet[f"{column}{row}"].row_height

    def get_column_width(self, column: str, row: int) -> float:
        """获取单元格宽度

        Args:
            column (str): 列, 用A、B、C等表示
            row (int): 行, 用1、2、3等表示(从1开始)

        Returns:
            float: 单元格宽度
        """
        return self.sheet[f"{column}{row}"].column_width

    def auto_fit(self, column: str, row: int, auto_save: bool=True):
        """自适应单元格

        Args:
            column (str): 列, 用A、B、C等表示
            row (int): 行, 用1、2、3等表示(从1开始)
            auto_save (bool, optional): 自动保存开关. Defaults to True.
        """
        self.sheet[f"{column}{row}"].autofit()
        if auto_save:
            self.book.save()
            
    def auto_fit_column(self, column: str, auto_save: bool=True):
        """自适应某一列

        Args:
            column (str): 列, 用A、B、C等表示
            auto_save (bool, optional): 自动保存开关. Defaults to True.
        """
        self.sheet[f"{column}1"].columns.autofit()
        if auto_save:
            self.book.save()
            
    def auto_fit_row(self, row: int, auto_save: bool=True):
        """自适应某一行

        Args:
            row (int): 行, 用1、2、3等表示(从1开始)
            auto_save (bool, optional): 自动保存开关. Defaults to True.
        """
        self.sheet[f"A{row}"].rows.autofit()
        if auto_save:
            self.book.save()
    
    def merge_cells(self, start_column: str, start_row: int, end_column: str, end_row: int):
        """合并单元格

        Args:
            start_column (str): 开始列, 用A、B、C等表示
            start_row (int): 开始行, 用1、2、3等表示(从1开始)
            end_column (str): 结束列, 用A、B、C等表示
            end_row (int): 结束行, 用1、2、3等表示(从1开始)
        """
        self.sheet.range(f"{start_column}{start_row}:{end_column}{end_row}").api.Merge()
    
    def unmerge_cells(self, start_column: str, start_row: int, end_column: str, end_row: int):
        """解除合并单元格

        Args:
            start_column (str): 开始列, 用A、B、C等表示
            start_row (int): 开始行, 用1、2、3等表示(从1开始)
            end_column (str): 结束列, 用A、B、C等表示
            end_row (int): 结束行, 用1、2、3等表示(从1开始)
        """
        self.sheet.range(f"{start_column}{start_row}:{end_column}{end_row}").api.UnMerge()
    
    def get_color(self, column: str, row: int) -> tuple:
        """获取单元格颜色

        Args:
            column (str): 列, 用A、B、C等表示
            row (int): 行, 用1、2、3等表示(从1开始)

        Returns:
            tuple: 单元格颜色, RGB值
        """
        return self.sheet[f"{column}{row}"].color

    def set_color(self, column: str, row: int, R: int, G: int, B: int, auto_save: bool=True):
        """填充单元格背景色

        Args:
            column (str): 列, 用A、B、C等表示
            row (int): 行, 用1、2、3等表示(从1开始)
            R (int): R值
            G (int): G值
            B (int): B值
            auto_save (bool, optional): 自动保存开关. Defaults to True.
        """
        self.sheet[f"{column}{row}"].color = (R, G, B)
        if auto_save:
            self.book.save()
            
    def set_formula(self, column: str, row: int, Formula: str, auto_save: bool=True):
        """设置某单元格的公式

        Args:
            column (str): 列, 用A、B、C等表示
            row (int): 行, 用1、2、3等表示(从1开始)
            Formula (str): excel公式
            auto_save (bool, optional): 自动保存开关. Defaults to True.
        """
        self.sheet[f"{column}{row}"].formula = Formula
        if auto_save:
            self.book.save()
            
    def get_formula(self, column: str, row: int):
        """获取某单元格的公式

        Args:
            column (str): 列, 用A、B、C等表示
            row (int): 行, 用1、2、3等表示(从1开始)

        Returns:
            单元格的公式
        """
        return self.sheet[f"{column}{row}"].formula

class ExcelUtils:
    """excel操作类, 若使用该工具类, 请使用如下命令安装第三方库: 
    \npip install openpyxl
    \npip install xlrd
    \npip install xlwt
    \npip install xlutils
    \npip install xlwings
    """
    DEFAULT_ENGINE_DICT = {".xlsx": "openpyxl", ".xls": "xlwings"}
    ENGINE_LIST = ["openpyxl", "xlutils", "xlwings"]

    @classmethod
    def get_sheet_list(cls, excel_path: str, engine: str=None, visible: bool=False) -> list[str]:    
        path = Path(excel_path)
        excel_path = str(path.resolve())
        if engine is None:
            engine = cls.DEFAULT_ENGINE_DICT[path.suffix]
        del path
        if engine == "openpyxl":
            book = load_workbook(excel_path, read_only=True)
            sheet_list = book.sheetnames
            book.close()
            del book
        elif engine == "xlutils":
            book = open_workbook(excel_path)
            sheet_list = book.sheet_names()
            del book
        elif engine == "xlwings":
            app = xw.App(visible=visible, add_book=False)
            book = app.books.open(excel_path)
            sheet_list = book.sheet_names
            book.close()
            app.quit()
            del app, book
        return sheet_list
        
    def __init__(self, excel_path: str, sheet_name: str="Sheet1", engine: str=None, encoding: str="utf-8", visible: bool=False):
        """初始化excel操作类

        Args:
            excel_path (str): excel路径
            sheet_name (str, optional): sheet名. Defaults to "Sheet1".
            engine (str, optional): 引擎名, 可以为 {openpyxl, xlutils, xlwings}, 若为None则自动选择引擎. Defaults to None.
            encoding (str, optional): 编码. Defaults to "utf-8".
            visible (bool, optional): 是否以可视化形式打开(只使用xlwings引擎有效). Defaults to False.
        """
        path = Path(excel_path)
        excel_path = str(path.resolve())
        """engine若为空, 则使用默认引擎"""
        if engine is None:
            engine = self.DEFAULT_ENGINE_DICT[path.suffix]
        print(f"Using engine: {engine}")
        
        """该Excel不存在, 则创建一个新的Excel"""
        if not path.exists():
            if engine == "openpyxl":
                from openpyxl import Workbook
                wb = Workbook()
                wb.create_sheet(sheet_name)
                wb.save(excel_path)
                del wb
            elif engine == "xlutils":
                from xlwt import Workbook
                wb = Workbook(encoding)
                wb.add_sheet(sheet_name)
                wb.save(excel_path)
                del wb
            elif engine == "xlwings":
                app = xw.App(visible=visible)
                book = app.books.add()
                book.sheets.add(sheet_name)
                book.save(excel_path)
                book.close()
                del app, book
        del path
        
        """通过引擎加载Excel"""
        if engine == "openpyxl":
            book = load_workbook(excel_path)
            sheet = book[sheet_name]
            self.__engine = OpenpyxlAdapter(book, sheet, excel_path, sheet_name)
        elif engine == "xlutils":
            book = open_workbook(excel_path)
            sheet = book.sheet_by_name(sheet_name)
            self.__engine = XlutilsAdapter(book, sheet, excel_path, sheet_name)
        elif engine == "xlwings":
            app = xw.App(visible=visible, add_book=False)
            book = app.books.open(excel_path)
            sheet = book.sheets[sheet_name]
            self.__engine = XlwingsAdapter(app, book, sheet, excel_path, sheet_name)
        else:
            raise ValueError(f"Unknow engine: {engine}, please select an engine in: {self.ENGINE_LIST}")
    
    def get_engine(self):
        return self.__engine

    def __enter__(self):
        return self.__engine

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.__engine:
            self.__engine.close()
            self.__engine = None
